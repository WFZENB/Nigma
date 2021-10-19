from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog
from PyQt5.QtGui import QFont
from PyQt5 import QtCore, QtWidgets
from qt_material import apply_stylesheet
from sys import argv, exit
from openpyxl import load_workbook

class Ui_MainWindow(object):
    def __init__(self, MainWindow):
        self.lbl_font = QFont() # Шрифт надписей
        self.btn_font = QFont() # Шрифт кнопок

        self.centralwidget      = QtWidgets.QWidget(MainWindow)             # Главный виджет
        self.gridLayout         = QtWidgets.QGridLayout(self.centralwidget) # Главный компоновщик

        self.verticalLayout     = QtWidgets.QVBoxLayout()                   # Вертикальный компоновщик
        self.horizontalLayout   = QtWidgets.QHBoxLayout()                   # Горизонтальный компоновщик
        self.label_init         = QtWidgets.QLabel(self.centralwidget)      # Надпись приветствия
        self.button_load        = QtWidgets.QPushButton(self.centralwidget) # Кнопка загрузки файла

        self.setupUi(MainWindow)

    def setupUi(self, MainWindow):
        # Выставление размеров шрифтов
        self.lbl_font.setPointSize(16)
        self.btn_font.setPointSize(14)

        # Определение размера окна и центрирование
        self.centralwidget.setMinimumSize(QtCore.QSize(900, 600))
        self.gridLayout.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)

        # Настройка надписи
        self.label_init.setMinimumSize(QtCore.QSize(300, 60))
        self.label_init.setFont(self.lbl_font)
        self.label_init.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
        self.verticalLayout.addWidget(self.label_init)

        # Добавление заполнителя
        spacerItem = QtWidgets.QSpacerItem(1, 40, QtWidgets.QSizePolicy.Expanding)
        self.verticalLayout.addItem(spacerItem)

        # Настройка кнопки
        self.button_load.setMinimumSize(QtCore.QSize(250, 60))
        self.button_load.setMaximumSize(QtCore.QSize(250, 60))
        self.button_load.setFont(self.btn_font)
        self.horizontalLayout.addWidget(self.button_load)
        self.verticalLayout.addLayout(self.horizontalLayout)

        # Финальная компоновка
        self.gridLayout.addLayout(self.verticalLayout, 0, 0)
        MainWindow.setCentralWidget(self.centralwidget)

        self.setTextUi()
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def setTextUi(self):
        # Выставление текста виджетам
        self.label_init.setText('Вас приветствует приложение анализа данных!\nПожалуйста, загрузите данные из файла')
        self.button_load.setText('Выбрать файл')


class Win(QMainWindow):
    def __init__(self):
        super(Win, self).__init__()
        self.ui = Ui_MainWindow(self)
        self.setWindowTitle('Analytics')
        self.init()

        self.ui.button_load.clicked.connect(self.show_dialog) # Бинд кнопки на вызов диалога

    def init(self):
        self.ui.button_load.setDefault(True)  # Установка фокуса на кнопку

    def show_dialog(self):
        dialog = QFileDialog.getOpenFileNames(self, "Open File", "", "XLM File (*.xlsx)")  # файл-диалог

        col_data = 1  # колонка с датой
        col_value = 2  # колонка с значением какого-либо параметра

        wb = load_workbook(filename=dialog[0][0], read_only=True, data_only=True)  # загрузка выбранного excel файла
        ws = wb.active  # обозначение главного элемента

        list_of_dates, list_of_value = [], []  # обозначаем списки с значениями дат и значениями параметра

        # указываем какие ячейки необходимо прочитать
        for i in ws.iter_rows(min_row=1, min_col=col_data, max_col=col_value, values_only=True):
            list_of_dates.append(i[0])  # добавляем значение в список дат
            list_of_value.append(i[1])  # добавляем значение в список значений

        # выводим значения в консоль
        for i in range(len(list_of_dates)):
            print(list_of_dates[i], "--", list_of_value[i])



def main():
    # Инициализация окна и запуск приложения
    app = QApplication(argv)
    apply_stylesheet(app, theme='dark_teal.xml')  # Установка темы приложения
    win = Win()
    win.show()
    exit(app.exec_())


if __name__ == '__main__':
    main()
